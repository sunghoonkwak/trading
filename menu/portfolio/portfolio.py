"""
This module handles portfolio menu management.
"""

def _check_portfolio_balance(merged_data, total_value_usd, current_weights, targets, exchange_rate):
    """
    Display weight differences with pagination UI.
    """
    from display import show_in_result_area, input_at
    from utils import get_fixed_width
    from data.data_service import get_weight_diffs

    if total_value_usd <= 0:
        display.add_alert("Total asset value is 0 or error.", "ERROR")
        return

    diffs, _ = get_weight_diffs()

    # Pagination
    import math
    page = 0
    page_size = 6
    total_pages = math.ceil(len(diffs) / page_size)
    if total_pages == 0: total_pages = 1

    while True:
        # Slice for current page
        start_idx = page * page_size
        end_idx = start_idx + page_size
        page_items = diffs[start_idx:end_idx]

        lines = []
        lines.append(f" [Portfolio Check] (Total: ${total_value_usd:,.0f}) - Page {page+1}/{total_pages}")

        # Header
        header = f" {'#':>2} | {'Ticker':<10} | {'Name':<30} | {'Curr %':>8} | {'Tgt %':>8} | {'Diff %':>9} | {'Qty':>6}"
        sep_len = len(header)

        lines.append("─" * sep_len)
        lines.append(header)
        lines.append("─" * sep_len)

        for i, item in enumerate(page_items):
            global_idx = start_idx + i + 1
            t = item["ticker"]

            # Truncate/Pad name for display (width 30)
            n = get_fixed_width(item['name'], 30)
            c_p = item["cur_w"] * 100
            t_p = item["tgt_w"] * 100
            d_p = item["diff"] * 100
            qty_diff = item["qty_diff"]

            # Qty string (int)
            qty_str = f"{int(qty_diff):+d}"

            # Colorize Diff (Positive = Buy = Green, Negative = Sell = Red)
            from display import COLOR_RED, COLOR_GREEN, COLOR_RESET
            symbol = "+" if d_p >= 0 else "-"

            c_str = f"{c_p:4.2f}%"
            t_str = f"{t_p:4.2f}%"
            base_d_str = f"{symbol}{abs(d_p):5.2f}%"

            # Color Logic
            if d_p > 0.5:
                colored_d_str = f"{COLOR_GREEN}{base_d_str}{COLOR_RESET}"
                colored_qt_str = f"{COLOR_GREEN}{qty_str:>6}{COLOR_RESET}"
            elif d_p < -0.5:
                colored_d_str = f"{COLOR_RED}{base_d_str}{COLOR_RESET}"
                colored_qt_str = f"{COLOR_RED}{qty_str:>6}{COLOR_RESET}"
            else:
                colored_d_str = base_d_str
                colored_qt_str = f"{qty_str:>6}"

            # Alignment padding for width 9 for Diff %
            target_width = 9
            padding = " " * (target_width - len(base_d_str))
            final_d_str = padding + colored_d_str

            lines.append(f" {global_idx:>2} | {t:<10} | {n} | {c_str:>8} | {t_str:>8} | {final_d_str} | {colored_qt_str}")

        remaining_lines = page_size - len(page_items)
        for _ in range(remaining_lines):
            lines.append("")  # Empty lines to keep UI stable

        lines.append("─" * sep_len)
        lines.append(" (Enter: Next Page, q: Return)")

        show_in_result_area(lines)

        # Input Loop
        cmd = input_at(13, 2, "Cmd (Ent/q): ").strip().lower()
        if cmd == 'q':
            break
        else:
            page = (page + 1) % total_pages


def _export_portfolio_excel(merged_data, current_weights, targets) -> bool:
    """
    Export portfolio to Excel (.xlsx) with formatted columns.
    Requires 'openpyxl' library.
    """
    from display import add_alert
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        add_alert("openpyxl not found. Please install: pip install openpyxl", "ERROR")
        return False

    from datetime import datetime
    import os

    rows = []

    for ticker, data in merged_data.items():
        qty = data["qty"]
        if qty <= 0: continue

        cur_price = data["cur_price"]
        total_inv = data["total_investment"]
        cur_val = data["current_value_native"]

        avg_price = total_inv / qty if qty > 0 else 0
        change = cur_val - total_inv
        # Return pct for Excel: Ratio (0.20 = 20%)
        ret_pct = (change / total_inv) if total_inv > 0 else 0

        currency = data["currency"]

        # Weigths
        cur_w = current_weights.get(ticker, 0.0)
        tgt_w = targets.get(ticker, 0.0)

        # Sort key logic
        sort_idx = 0 if currency == "USD" else 1
        if data.get("type") == "CASH": sort_idx += 2

        rows.append({
            "ticker": ticker,
            "name": data["name"],
            "cur_w": cur_w,
            "tgt_w": tgt_w,
            "qty": qty,
            "avg_price": avg_price,
            "cur_price": cur_price,
            "total_inv": total_inv,
            "cur_val": cur_val,
            "change": change,
            "ret_pct": ret_pct,
            "sort_key": (sort_idx, ticker)
        })

    # Sort
    rows.sort(key=lambda x: x["sort_key"])

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # Headers
    headers = ["Ticker", "Name", "Cur Wgt", "Tgt Wgt", "Qty", "Avg Price", "Cur Price", "Invest", "Cur Val", "Change", "Ret %"]
    ws.append(headers)

    # Style Headers: Bold + Center
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align

    # Write Data
    for r in rows:
        ws.append([
            r["ticker"], r["name"], r["cur_w"], r["tgt_w"], r["qty"],
            r["avg_price"], r["cur_price"], r["total_inv"], r["cur_val"],
            r["change"], r["ret_pct"]
        ])

    # Formatting
    # Columns (1-based):
    # A(1):Ticker, B(2):Name, C(3):CurW, D(4):TgtW, E(5):Qty
    # F(6):Avg, G(7):CurP, H(8):Inv, I(9):Val, J(10):Chg, K(11):Ret%

    number_fmt = "#,##0.00"
    pct_fmt = "0.00%"

    for row_idx in range(2, len(rows) + 2):
        # Weights (C, D)
        ws.cell(row=row_idx, column=3).number_format = pct_fmt
        ws.cell(row=row_idx, column=4).number_format = pct_fmt

        # Qty (E)
        ws.cell(row=row_idx, column=5).number_format = "#,##0.00"

        # Prices/Values (F-J)
        for col_idx in range(6, 11):
            ws.cell(row=row_idx, column=col_idx).number_format = number_fmt

        # Return (K)
        ws.cell(row=row_idx, column=11).number_format = pct_fmt

    # Column Widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 10

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"portfolio_export_{timestamp}.xlsx"

    # Path resolution: trading/exports/
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) # trading/
    export_dir = os.path.join(base_dir, "exports")

    if not os.path.exists(export_dir):
        os.makedirs(export_dir)

    filepath = os.path.join(export_dir, filename)

    try:
        wb.save(filepath)
        add_alert(f"Excel exported: {filename}", "SUCCESS")
        return True
    except Exception as e:
        add_alert(f"Excel save failed: {e}", "ERROR")
        return False


def _print_portfolio_summary(stats, exchange_rate, total_value_usd):
    """
    Print the portfolio summary table.
    """
    from display import show_in_result_area

    # Build summary lines
    lines = []
    lines.append(f" [Portfolio] (Rate: {exchange_rate:,.2f} KRW/USD) (Total: ${total_value_usd:,.0f})")
    lines.append(f" {'':10} │ {'USD':^27} │ {'KRW':^27} │ {'%':^6}")
    lines.append("─" * 80)

    def _align_pair(val1, val2, curr1, curr2, width=6):
        """Format and align two currency values for a table cell."""
        sym1 = "$ " if curr1 == "USD" else "₩ "
        sym2 = "$ " if curr2 == "USD" else "₩ "

        def _get_num(v, c):
            if c == "USD":
                return f"{v/1000:,.1f}K" if abs(v) >= 1000 else f"{v:,.2f}"
            else:
                if abs(v) >= 1000000: return f"{v/1000000:,.1f}M"
                return f"{v/1000:,.0f}K" if abs(v) >= 1000 else f"{v:,.0f}"

        s1 = f"{sym1}{_get_num(val1, curr1).rjust(width)}"
        s2 = f"{sym2}{_get_num(val2, curr2).rjust(width)}"
        return f"{s1} / {s2}"

    # US Assets row
    us_usd = _align_pair(stats['us_stock_usd'], stats['us_cash_usd'], "USD", "USD")
    us_krw = _align_pair(stats['us_stock_krw'], stats['us_cash_krw'], "KRW", "KRW")
    lines.append(f" {'US Assets':10} │ {us_usd:^27} │ {us_krw:^27} │ {stats['us_pct']:5.1f}%")

    # KR Assets row
    kr_usd = _align_pair(stats['kr_stock_usd'], stats['kr_cash_usd'], "USD", "USD")
    kr_krw = _align_pair(stats['kr_stock_krw'], stats['kr_cash_krw'], "KRW", "KRW")
    lines.append(f" {'KR Assets':10} │ {kr_usd:^27} │ {kr_krw:^27} │ {stats['kr_pct']:5.1f}%")

    lines.append("─" * 80)

    # Total row
    tot_usd = _align_pair(stats['total_stock_usd'], stats['total_cash_usd'], "USD", "USD")
    tot_krw = _align_pair(stats['total_stock_krw'], stats['total_cash_krw'], "KRW", "KRW")
    lines.append(f" {'Total':10} │ {tot_usd:^27} │ {tot_krw:^27} │")

    # Cash ratio row
    us_cash_str = f"{stats['us_cash_ratio']:>14.1f}        %"
    kr_cash_str = f"{stats['kr_cash_ratio']:>14.1f}        %"
    lines.append(f" {'Cash':10} │ {us_cash_str:27} │ {kr_cash_str:27} │")
    lines.append("─" * 80)
    lines.append(" 1. Check Portfolio  2. Excel Export  3. Value Averaging  q. Exit")

    show_in_result_area(lines)


def portfolio_menu():
    """
    Portfolio menu interface using modular get_portfolio() function.
    Displays summary and handles menu interactions.
    """
    from display import show_in_result_area, input_at
    from data.data_service import get_portfolio_data

    # Get portfolio data
    portfolio_data = get_portfolio_data()

    if portfolio_data.get("error"):
        display.add_alert(f"Portfolio error: {portfolio_data['error']}", "ERROR")
        return

    # Extract data
    merged_data = portfolio_data["merged_data"]
    total_value_usd = portfolio_data["total_value_usd"]
    current_weights = portfolio_data["current_weights"]
    targets = portfolio_data["targets"]
    stats = portfolio_data["stats"]
    exchange_rate = portfolio_data["exchange_rate"]

    while True:
        _print_portfolio_summary(stats, exchange_rate, total_value_usd)

        choice = input_at(12, 2, "Enter Choice: ").strip().lower()

        if choice == '1':
            _check_portfolio_balance(merged_data, total_value_usd, current_weights, targets, exchange_rate)
            input_at(12, 2, "Press Enter to continue...")
        elif choice == '2':
            _export_portfolio_excel(merged_data, current_weights, targets)
            input_at(12, 2, "Press Enter to continue...")
        elif choice == '3':
            from . import value_averaging
            from display import show_in_result_area, input_at

            # Call refactored get_daily_report (returns multi-strategy results)
            res = value_averaging.get_daily_report()

            # Build Display Lines
            lines = []
            date = res.get("date", "N/A")
            results = res.get("results", [])
            total_orders = res.get("total_orders", [])

            # Header
            lines.append(f" [Value Averaging] {date} ET")
            lines.append("=" * 50)

            if res.get("error"):
                lines.append(f" ERROR: {res.get('error')}")
            elif not results:
                lines.append(" No strategies configured.")
            else:
                for r in results:
                    if r.get("error"):
                        lines.append(f" {r.get('target_ticker', 'Unknown')}: {r['error']}")
                        continue

                    target_ticker = r.get("target_ticker", "N/A")
                    already_executed = r.get("already_executed", False)
                    curr_p = r.get("current_price", 0)
                    daily_b = r.get("daily_budget", 0)
                    buy_amt = r.get("daily_target_amount", 0)
                    day_count = r.get("day_count", 0)
                    orders = r.get("orders", [])

                    status_str = "✓" if already_executed else " "
                    lines.append(f" {status_str} {target_ticker} | Day: {day_count} | Price: ${curr_p:,.2f}")
                    lines.append(f"   Budget: ${daily_b:,.2f} | Buy: ${buy_amt:,.2f}")

                    if already_executed:
                        lines.append("   [Executed today]")
                        executed_orders = r.get("executed_orders", [])
                        for o in executed_orders:
                            o_qty = o.get('qty', 0)
                            o_price = o.get('price', 0)
                            o_type = o.get('order_type', 'LOC')
                            lines.append(f"   > {o_qty} qty @ ${o_price:.2f} ({o_type})")
                    elif curr_p == 0:
                        lines.append("   [Price unavailable]")
                    elif orders:
                        for o in orders:
                            lines.append(f"   > {o['qty']} qty @ ${o['price']:.2f} (LOC)")
                    else:
                        lines.append("   [No order needed]")
                    lines.append("-" * 50)

            show_in_result_area(lines)

            # Execution Prompt (only if has orders)
            if total_orders:
                from utils import is_market_holiday
                if is_market_holiday("NYSE"):
                    lines.append(" 🚫 휴장일 - 주문 불가")
                    show_in_result_area(lines)
                else:
                    confirm = input_at(13, 2, f" Execute {len(total_orders)} order(s)? (y/n): ").strip().lower()
                    if confirm == 'y':
                        exec_res = value_averaging.execute_orders(res)
                        result_lines = [lines[0], lines[1]]  # Keep header
                        result_lines.append(" Execution Result:")
                        for r in exec_res:
                            ticker = r.get('ticker', 'Unknown')
                            if r.get('skipped'):
                                result_lines.append(f"  ⏭️ {ticker}: {r.get('message', 'Skipped')}")
                            else:
                                status = "✓" if r.get('success') else "✗"
                                result_lines.append(f"  {status} {ticker}: {r.get('message', 'Unknown')}")
                        show_in_result_area(result_lines)

            input_at(12, 2, " Press Enter to continue...")
        elif choice == 'q':
            break
